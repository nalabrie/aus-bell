#!/usr/bin/env python

import fnmatch
import logging
import pickle
from datetime import datetime, timedelta
from itertools import cycle, zip_longest
from os import chdir, mkdir, listdir, remove
from pathlib import Path
from platform import system
from random import shuffle
from subprocess import Popen, run
from sys import exit
from time import sleep

import coloredlogs
import yaml
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException as openpyxl_InvalidFileException
from yt_dlp import YoutubeDL


# ---- CLASSES ----

class DummyLogger:
    # This class is needed to disable the output from yt-dlp
    def debug(self, msg):
        pass

    def info(self, msg):
        pass

    def warning(self, msg):
        pass

    def error(self, msg):
        pass


# ---- FUNCTIONS ----

def show_intro():
    """
    Displays the intro message to the terminal at the start of the script.
    """
    message = '''
        ---- Usage Instructions ----

Press "Ctrl + C" to play the next bell immediately.
Press "Ctrl + C" to stop playing the bell immediately.
Close the terminal window to stop this script at any time.
'''
    print(message)


def show_version():
    """
    Outputs the script's version to the LOGGER.
    """
    try:
        with open("VERSION", mode='r') as f:
            LOGGER.info(f"aus-bell version {f.read(7)}")
    except FileNotFoundError:
        LOGGER.error('"VERSION" file is missing, continuing anyway')


def show_bell_schedule():
    """
    Outputs the bell schedule to the LOGGER.
    """
    LOGGER.debug("Bell schedule:")
    LOGGER.debug("--------------")
    for time in BELL_SCHEDULE:
        LOGGER.debug(time.strftime('%I:%M %p'))


def play_media(file_path):
    """
    Plays audio files using "ffplay". Can be stopped early with "Ctrl+C".
    :param file_path: Path to the audio file
    """
    try:
        run(f"{FFPLAY_PATH} -nodisp -autoexit -loglevel fatal {file_path}")
    except KeyboardInterrupt:
        LOGGER.warning(f"User requested to stop the bell early at {datetime.now().strftime('%I:%M %p')}")


def set_play_order():
    """
    Sets the order that the bell will play.
    """
    # for each file in media folder...
    for file in fnmatch.filter(listdir(), 'bell_*.mkv'):
        PLAYLIST.append(file)
    shuffle(PLAYLIST)
    global PLAY_CYCLE
    PLAY_CYCLE = cycle(PLAYLIST)
    LOGGER.info(f"playlist with {len(PLAYLIST)} songs has been shuffled")


def ring_bell():
    """
    Rings the next bell in the playlist.
    """
    song = next(PLAY_CYCLE)
    LOGGER.info(f"playing file: {song}")
    play_media(song)


def sleep_until(target: datetime):
    """
    Sleeps script until the target date and time. Pressing "Ctrl+C" will cancel the sleep.
    :param target: date and time to sleep until (must be a "datetime" object)
    """
    now = datetime.now()
    delta = target - now

    if delta > timedelta(0):
        LOGGER.info(f"Waiting until next bell at {target.strftime('%I:%M %p')}")
        try:
            sleep(delta.total_seconds())
        except KeyboardInterrupt:
            LOGGER.warning(f"User requested to play the bell early at {datetime.now().strftime('%I:%M %p')}")
    else:
        raise ValueError('"sleep_until()" cannot sleep a negative amount of time.')


def create_bell_schedule():
    """
    Initiates a list with the current day's bell schedule.
    """
    today = datetime.now().replace(second=0, microsecond=0)
    for time in CFG_DICT["bell_schedule"]:
        BELL_SCHEDULE.append(today.replace(hour=time["hour"], minute=time["minute"]))


def setup_dirs():
    """
    Prepare needed directories and change working directory to the "media" folder.
    """
    try:
        mkdir("media")
        LOGGER.info('"media" directory was created')
    except FileExistsError:
        pass
    chdir("media")


def setup_dynamic_paths():
    """
    Prepare needed file paths.
    These paths can change between systems.
    """
    global LINKS_PATH, LOG_PATH
    try:
        LINKS_PATH = CFG_DICT["links_spreadsheet_path"]
    except KeyError:
        LOGGER.critical('"config.yaml" file does not contain the key "links_spreadsheet_path". Stopping now.')
        input("\nPress ENTER to exit")
        exit(1)
    try:
        LOG_PATH = CFG_DICT["log_file_path"]
    except KeyError:
        LOGGER.critical('"config.yaml" file does not contain the key "log_file_path". Stopping now.')
        input("\nPress ENTER to exit")
        exit(1)


def setup_static_paths():
    """
    Prepare needed file paths.
    These paths are the same on every system.
    """
    global CACHE_PATH, FFMPEG_PATH, FFPLAY_PATH, CFG_PATH

    CACHE_PATH = str(Path("cache.dat").resolve())  # it is ok if this does not exist
    if system() == "Windows":
        try:
            FFMPEG_PATH = str(Path("ffmpeg.exe").resolve(strict=True))
            FFPLAY_PATH = str(Path("ffplay.exe").resolve(strict=True))
        except FileNotFoundError:
            LOGGER.critical('Missing ffmpeg and/or ffplay. Run "setup.bat" to download both. Stopping now.')
            input("\nPress ENTER to exit")
            exit(1)
    else:
        # this situation should never occur because only Windows is supported (for now)
        try:
            FFMPEG_PATH = str(Path("ffmpeg").resolve(strict=True))
            FFPLAY_PATH = str(Path("ffplay").resolve(strict=True))
        except FileNotFoundError:
            LOGGER.critical("placeholder")
            input("\nPress ENTER to exit")
            exit(1)
    CFG_PATH = str(Path("config.yaml").resolve(strict=True))


def set_current_media_list():
    """
    Fills out the global "CURRENT_MEDIA_LIST_NUMBERS" with "bell numbers".
    These numbers are pulled directly from the file names of already downloaded media.
    """
    for file in fnmatch.filter(listdir(), 'bell_*.mkv'):
        num_string = ""
        for char in file:
            if char.isdigit():
                num_string += char
        CURRENT_MEDIA_LIST_NUMBERS.append(int(num_string))
    CURRENT_MEDIA_LIST_NUMBERS.sort()


def setup_logging():
    """
    Sets up the config for the LOGGER.
    """
    log_format = "%(asctime)s %(name)s %(levelname)s %(message)s"

    # create a "FileHandler" object
    fh = logging.FileHandler(LOG_PATH, mode='w')
    fh.setLevel(logging.DEBUG)

    # create a "ColoredFormatter" to use as formatter for the FileHandler
    formatter = coloredlogs.ColoredFormatter(log_format)
    fh.setFormatter(formatter)
    LOGGER.addHandler(fh)

    # install the "coloredlogs" module
    coloredlogs.install(level="DEBUG", logger=LOGGER, fmt=log_format)

    # show all paths
    LOGGER.info("Script started, logging initiated")
    LOGGER.info(f'Path to links spreadsheet: "{LINKS_PATH}"')
    LOGGER.info(f'Path to log file: "{LOG_PATH}"')
    LOGGER.info(f'Path to cache file: "{CACHE_PATH}"')
    LOGGER.info(f'Path to config file: "{CFG_PATH}"')
    LOGGER.info(f'Path to ffmpeg executable: "{FFMPEG_PATH}"')
    LOGGER.info(f'Path to ffplay executable: "{FFPLAY_PATH}"')


def load_prev_urls():
    """
    Loads the URL list from the last time this script was ran.
    This is needed to compare if changes/deletions to the spreadsheet were made.
    """
    global PREV_URLS
    try:
        with open(CACHE_PATH, "rb") as f:
            PREV_URLS = pickle.load(f)
        LOGGER.info("Loaded list of URLs from the previous run")
    except FileNotFoundError:
        # continuing without a cache file is ok, but show warning
        LOGGER.warning("No cache file found, cannot load previous list of URLs")


def save_curr_urls():
    """
    Saves the current URL list to a cache file.
    Used in conjunction with "load_prev_urls()"
    """
    with open(CACHE_PATH, "wb") as f:
        pickle.dump(ALL_URLS, f, pickle.HIGHEST_PROTOCOL)


def compare_urls():
    """
    Compares current list of URLs with the previous run's URL list.
    This determines which new links need downloaded and which need deleted/overwritten.
    IMPORTANT: make sure to call "delete_unused_media()" right after this function.
    """
    for i, (new, old) in enumerate(zip_longest(ALL_URLS, PREV_URLS)):
        if new == old:
            # this file is already downloaded, skip it
            continue
        else:
            if new is None:
                # this link was removed from the spreadsheet and not replaced, queue for deletion
                TO_BE_DELETED_MEDIA_LIST_NUMBERS.append(i)
            else:
                # this file is needed, queue for download
                NEEDED_MEDIA_LIST_NUMBERS.append(i)
                if i in CURRENT_MEDIA_LIST_NUMBERS:
                    # new link overwrites old media file, queue old file for deletion
                    # DELETE MUST OCCUR RIGHT AFTER THIS FUNCTION (because of this situation)
                    TO_BE_DELETED_MEDIA_LIST_NUMBERS.append(i)


def delete_unused_media():
    """
    Deletes media files that are no longer listed in the spreadsheet.
    If a file doesn't exist, it is simply skipped.
    The "CURRENT_MEDIA_LIST_NUMBERS" variable is refreshed afterward.
    """
    # delete all unused media files
    for i in TO_BE_DELETED_MEDIA_LIST_NUMBERS:
        file = f"bell_{i}.mkv"
        try:
            remove(file)
            LOGGER.debug(f'deleted file "{file}"')
        except FileNotFoundError:
            LOGGER.warning(f'Cannot delete file "{file}" because it does not exist. Skipping deletion.')

    # refresh current media list
    CURRENT_MEDIA_LIST_NUMBERS.clear()
    set_current_media_list()


def read_url_file():
    """
    Read list of URLs from "links.xlsx" spreadsheet and store them in the global "ALL_URLS" variable.
    """
    try:
        wb = load_workbook(filename=LINKS_PATH, read_only=True)
        ws = wb["Sheet1"]
        for count, row in enumerate(ws.iter_rows(max_col=1, values_only=True)):
            link = row[0]
            ALL_URLS.append(link)
    except FileNotFoundError:
        LOGGER.critical(f'"{LINKS_PATH}" does not exist, stopping now')
        input("\nPress ENTER to exit")
        exit(1)
    except PermissionError:
        LOGGER.critical(f'Permission was denied to read spreadsheet file (located at: "{LINKS_PATH}"). '
                        'The file is likely open in Excel. Close Excel and run this script again. Stopping now.')
        input("\nPress ENTER to exit")
        exit(1)
    except openpyxl_InvalidFileException:
        LOGGER.critical("Cannot not open links spreadsheet, it is an invalid file type. "
                        "Supported formats are: .xlsx, .xlsm, .xltx, .xltm. Stopping now.")
        input("\nPress ENTER to exit")
        exit(1)


def download_all():
    """
    Downloads all media from URLs in list with yt-dlp,
    converts first "MEDIA_LENGTH" minute(s) to a mkv file (copies original codec)
    """
    # ---- extract all audio links from all valid links ----
    with YoutubeDL(OPTS) as ydl:
        extracted_urls = []
        out_file_names = []
        LOGGER.info(f"extracting audio URLs from {len(NEEDED_MEDIA_LIST_NUMBERS)} links with yt-dlp")
        for file_num in NEEDED_MEDIA_LIST_NUMBERS:
            # extract the audio track URL from each link
            try:
                extracted_urls.append(ydl.extract_info(ALL_URLS[file_num], download=False)["url"])
                out_file_names.append(f"bell_{file_num}.mkv")
                LOGGER.info(f"yt-dlp extracted the audio URL for {ALL_URLS[file_num]}")
            except (TypeError, KeyError):
                # tried to download an invalid link, just skip this one
                LOGGER.warning(f"yt-dlp skipped an invalid URL: {ALL_URLS[file_num]}")
                extracted_urls.append(None)
                out_file_names.append(None)
                continue

    # ---- download all valid audio files with ffmpeg ----
    download_count = 0
    for link in extracted_urls:
        if link is not None:
            # count all valid links
            download_count += 1
    LOGGER.info(f"Downloading {download_count} files with ffmpeg")
    ffmpeg_processes = []
    for file_num, link, file_name in zip(NEEDED_MEDIA_LIST_NUMBERS, extracted_urls, out_file_names):
        if link is None:
            # skip invalid link
            continue
        # using ffmpeg:
        #   1. download all media at the same time
        #   2. convert first "MEDIA_LENGTH" minute(s) to a mkv file (copy audio codec)
        #   3. output file names are sequential (they match the order in spreadsheet)
        #   4. wait to return until all ffmpeg instances are finished
        ffmpeg_processes.append(
            Popen(
                f"{FFMPEG_PATH} -loglevel fatal -n -ss 00:00:00 -to 00:{MEDIA_LENGTH} -i {link} -vn -c:a copy {file_name}"
            )
        )
    for process in ffmpeg_processes:
        # wait for all downloads to finish
        while process.poll() is None:
            sleep(0.5)
        else:
            LOGGER.info(f"ffmpeg PID {process.pid} is finished")


def load_config():
    """
    Loads all data from the config file "config.yaml".
    """
    global CFG_DICT
    try:
        with open(CFG_PATH, 'r') as f:
            CFG_DICT = yaml.safe_load(f)
    except FileNotFoundError:
        LOGGER.critical(
            'file "config.yaml" does not exist in root directory of program. Cannot load config. Stopping now.'
        )
        input("\nPress ENTER to exit")
        exit(1)


def main():
    """
    Main program routine. Runs when script is executed.
    """
    setup_static_paths()
    load_config()
    setup_dynamic_paths()
    show_intro()
    setup_logging()
    show_version()
    setup_dirs()
    set_current_media_list()
    create_bell_schedule()
    show_bell_schedule()
    load_prev_urls()
    read_url_file()
    save_curr_urls()
    compare_urls()
    delete_unused_media()
    download_all()
    set_play_order()
    for time in BELL_SCHEDULE:
        try:
            sleep_until(time)
        except ValueError:
            # bell has already happened, so skip it
            continue
        ring_bell()
    LOGGER.debug("Script finished cleanly")
    input("\nPress ENTER to exit")


# ---- GLOBALS ----

# all global variables needed for "main()"
MEDIA_LENGTH = "03:00"  # max length to trim downloaded media in MM:SS format (string)
CURRENT_MEDIA_LIST_NUMBERS = []  # list of numbers representing available media files
NEEDED_MEDIA_LIST_NUMBERS = []  # list of numbers representing the needed media files
TO_BE_DELETED_MEDIA_LIST_NUMBERS = []  # list of numbers representing which media files need deleted
BELL_SCHEDULE = []  # bell schedule list
PREV_URLS = []  # list of URLS from the previous run
ALL_URLS = []  # all URLs that are in the spreadsheet (valid or not!)
LINKS_PATH = None  # file path to "links.xlsx" (where media URLs are stored)
LOG_PATH = ""  # file path to "bell.log" (where the logger saves to)
LOGGER = logging.getLogger("aus-bell")  # initiate logger
CACHE_PATH = ""  # file path to "cache.dat" (where the previous run's URLs are stored)
FFMPEG_PATH = ""  # path to ffmpeg executable
FFPLAY_PATH = ""  # path to ffplay executable
PLAYLIST = []  # bell play order
PLAY_CYCLE = cycle([])  # circular list version of "PLAYLIST"
CFG_DICT = {}  # dictionary of data read from "config.yaml"
CFG_PATH = ""  # path to the config file (config.yaml)
OPTS = {  # yt-dlp arguments
    'format': 'bestaudio/best',
    'ignoreerrors': True,
    "quiet": True,
    "noprogress": True,
    "no_warnings": True,
    "logger": DummyLogger()
}

if __name__ == '__main__':
    main()
